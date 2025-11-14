import pandas as pd
import numpy as np
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import matplotlib.pyplot as plt
import seaborn as sns
def read_final_data():
    final_data_file_path = "../Merge_Data/final.csv"
    df = pd.read_csv(final_data_file_path, index_col=0)
    return df

def check_missing(df: pd.DataFrame, exclude_cols: list = None, target_cols: list = None):
    key_cols = ['iso_code', 'year', 'country']
    if exclude_cols is not None:
        key_cols.extend(exclude_cols)
    
    if target_cols is not None:
        missing_df = df[df[target_cols].isna().all(axis=1)]
        return missing_df
    
    feature_cols = [c for c in df.columns if c not in key_cols]
    missing_df = df[df[feature_cols].isna().all(axis=1)]

    return missing_df

def remove_missing_ft(df: pd.DataFrame, exclude_cols: list = None, target_cols: list = None):
    key_cols = ['iso_code', 'year', 'country']
    if exclude_cols is not None:
        key_cols.extend(exclude_cols)

    if target_cols is not None:
        mask = df[target_cols].isna().all(axis=1)
        df_clean = df[~mask].copy()
        return df_clean
    
    feature_cols = [c for c in df.columns if c not in key_cols]
    mask = df[feature_cols].isna().all(axis=1)
    df_clean = df[~mask].copy()
    return df_clean

def remove_column(df: pd.DataFrame, cols_to_remove: list = None, inplace: bool = False):
    existing_cols = [c for c in cols_to_remove if c in df.columns]
    missing_cols = [c for c in cols_to_remove if c not in df.columns]
    if inplace:
        df.drop(columns=existing_cols, inplace=True)
        return None
    else:
        return df.drop(columns=existing_cols)

def fill_missing_by_country(df: pd.DataFrame):
    df_filled = df.copy()

    for col in df.columns:
        if col in ['iso_code', 'year', 'country']:
            continue

        if col in ['ft_tax', 'ft_electriccarssold', 'ft_nonelectriccarsales']:
            df_filled[col] = df_filled[col].fillna(0)
        elif col in ['ft_area_ha']:
            df_filled[col] = (
                df_filled.groupby('iso_code')[col]
                .ffill().bfill()
            )
        elif col in ['ft_hdi']:
            df_filled[col] = (
                df_filled.groupby('iso_code')[col]
                .apply(lambda s: s.interpolate(method='linear')).ffill().bfill()
            )

        elif col in ['ft_fossil_fuel']:
            df_filled[col] = (
                df_filled.groupby('iso_code')[col]
                .apply(lambda s: s.fillna(s.mean())).ffill().bfill()
            )
        else:
            df_filled[col] = df_filled[col].ffill().bfill()

    return df_filled


def preprocess_data(df: pd.DataFrame):
    df = remove_missing_ft(df=df, exclude_cols=['ft_co2']) # remove record bi missing o tat ca feature (ngoai tru co2)    
    df = remove_missing_ft(df=df, target_cols=['ft_co2']) # remove record bi missing co2

    return df

def raw_eda(df: pd.DataFrame, output_file : str = "raw_eda_result.xlsx"):

    df_info = pd.DataFrame({
        "column": df.columns.to_list(),
        "dtype": df.dtypes,
        "missing": df.isna().sum(),
        "missing_pct": (df.isna().mean() * 100).round(2),
        "unique": df.nunique()
    })
    info_explain = {
        "Metric": df_info.columns.tolist(),
        "Description": [
            "Tên của các cột",
            "Kiểu dữ liệu của cột", 
            "Số lượng missing value(NaN)",
            "Tỉ lệ % missing value",
            "Số lượng giá trị duy nhất",
        ],
        "Unit": [
            "-",
            "-",
            "count",
            "%",
            "count"
        ]
    }
    df_info_explain = pd.DataFrame(info_explain)

    df_describe = df.describe().T.reset_index().rename(columns={"index": "column"})

    df_corr = df.corr(numeric_only=True)

    missing_all_feature = check_missing(df=df, exclude_cols=["ft_co2"])
    missing_co2 = check_missing(df=df, target_cols=["ft_co2"])
    df_summary = pd.DataFrame({
        "num_rows": [df.shape[0]],
        "num_cols": [df.shape[1]],
        "missing_all_feature": [len(missing_all_feature)],
        "missing_co2": [len(missing_co2)]
    }).T.reset_index()
    df_summary.columns = ["Metric", "Value"]

    with pd.ExcelWriter(output_file) as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)

        df_info.to_excel(writer, sheet_name="Info", index=False, startcol=0, startrow=0)
        df_info_explain.to_excel(writer, sheet_name="Info", index=False, startcol=len(df_info.columns) + 2, startrow= 2)

        df_describe.to_excel(writer, sheet_name="Describe", index=False)
        df_corr.to_excel(writer, sheet_name="Correlation")

        workbook = writer.book

        font = Font(name="Calibri", size=14, bold=False)

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            for row in worksheet.iter_rows():
                for cell in row:
                    cell.font = font

            for col in worksheet.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                worksheet.column_dimensions[col_letter].width = max_length + 2
        
        worksheet_info = workbook["Info"]
        header_row = 1
        missing_pct_col = None
        for cell in worksheet_info[header_row]:
            if (str(cell.value).strip().lower() == "missing_pct"):
                missing_pct_col = cell.column_letter
                break
        
        if missing_pct_col:
            start_row = header_row + 1
            end_row = worksheet_info.max_row
            cell_range = f"{missing_pct_col}{start_row}:{missing_pct_col}{end_row}"

            color_scale = ColorScaleRule(
                start_type="num", start_value=0, start_color="63BE7B",
                mid_type="num", mid_value=50, mid_color="FFEB84",
                end_type="num", end_value=100, end_color="F8696B"
            )

            worksheet_info.conditional_formatting.add(cell_range, color_scale)

    

def main():
    df = read_final_data()
    raw_eda(df=df)
    df = preprocess_data(df=df)
    raw_eda(df=df, output_file="preprocess_result.xlsx")
    print(df.isna().mean().sort_values(ascending=False))

if __name__ == "__main__":
    main()